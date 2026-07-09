# -*- coding: utf-8 -*-
"""
過去提出済み提案書（PDF / Excel）を管理し、AI生成用のコンテキストを提供するモジュール。
提案書が蓄積されるほど精度が上がる「進化型ナレッジベース」として機能する。
"""
import os, json, re
from datetime import datetime
from collections import Counter

_BASE    = os.path.dirname(__file__)
REF_DIR  = os.path.join(_BASE, "data", "references")
REF_META = os.path.join(REF_DIR, "metadata.json")


def _init():
    os.makedirs(REF_DIR, exist_ok=True)
    if not os.path.exists(REF_META):
        with open(REF_META, "w", encoding="utf-8") as f:
            json.dump([], f)


def _load_meta() -> list:
    _init()
    try:
        with open(REF_META, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        import shutil
        bak = REF_META + ".bak"
        try:
            shutil.copy2(REF_META, bak)
        except Exception:
            pass
        return []


def _save_meta(meta: list):
    tmp = REF_META + ".tmp"
    with open(tmp, "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False, indent=2)
    os.replace(tmp, REF_META)


# ──────────────────────────────────────────
# テキスト抽出
# ──────────────────────────────────────────
def _extract_pdf(path: str) -> str:
    try:
        from pypdf import PdfReader
        r = PdfReader(path)
        return "\n".join(p.extract_text() or "" for p in r.pages).strip()
    except Exception as e:
        return f"[PDF抽出エラー: {e}]"


def _extract_excel(path: str) -> str:
    try:
        import openpyxl, warnings
        warnings.filterwarnings("ignore")
        wb = openpyxl.load_workbook(path, data_only=True)
        ws = None
        for name in wb.sheetnames:
            if "様式" in name or "提案" in name:
                ws = wb[name]
                break
        if ws is None:
            ws = wb.active

        texts = []
        for row in ws.iter_rows():
            for cell in row:
                v = cell.value
                if isinstance(v, str) and len(v.strip()) >= 8:
                    texts.append(v.strip())
        return "\n".join(texts)
    except Exception as e:
        return f"[Excel抽出エラー: {e}]"


def extract_text(file_path: str, filename: str) -> str:
    ext = os.path.splitext(filename)[1].lower()
    if ext == ".pdf":
        return _extract_pdf(file_path)
    elif ext in (".xlsx", ".xls"):
        return _extract_excel(file_path)
    return ""


# ──────────────────────────────────────────
# テキスト解析
# ──────────────────────────────────────────
def _clean_line(line: str) -> str:
    """PDF抽出時のノイズ（ページ番号・余分な数字・記号）を除去する"""
    # 末尾の半角数字・スペース・改ページ記号を除去
    line = re.sub(r'\s+\d{1,3}\s*$', '', line)
    # 連続スペースを1つに
    line = re.sub(r'　+', '　', line)
    return line.strip()


def _extract_chui_lines(text: str) -> list:
    """留意点テキストの行を抽出（ラベル除去・ノイズ除去済み）"""
    results = []
    for line in text.splitlines():
        line = _clean_line(line)
        if not line:
            continue
        if re.search(r'留意点[①②③]', line) or ('留意する' in line and len(line) > 10):
            body = re.sub(r'^.{0,8}留意点[①②③０-９\d]?\s*[:：]?\s*', '', line).strip()
            body = _clean_line(body)
            if len(body) >= 10 and '留意する' in body:
                results.append(body)
    return results


def _extract_riyu_lines(text: str) -> list:
    """理由テキストの行を抽出（ラベル除去・ノイズ除去済み）"""
    results = []
    for line in text.splitlines():
        line = _clean_line(line)
        if not line:
            continue
        if re.search(r'^.{0,8}理由[①②③④]', line) and 'ため' in line:
            body = re.sub(r'^.{0,8}理由[①②③④]\s*[:：]?\s*', '', line).strip()
            body = _clean_line(body)
            if len(body) >= 10:
                results.append(body)
    return results


def _extract_paired_blocks(text: str) -> list:
    """
    留意点＋その理由①〜④のペアブロックを抽出する。
    Returns: [{"留意点": str, "理由": [str, ...]}, ...]
    """
    blocks = []
    lines = [l.strip() for l in text.splitlines() if l.strip()]

    i = 0
    while i < len(lines):
        line = _clean_line(lines[i])
        # 留意点行の検出
        is_chui = (re.search(r'留意点[①②③]', line) or
                   ('留意する' in line and len(line) > 10 and not re.search(r'^.{0,4}理由', line)))
        if not is_chui:
            i += 1
            continue

        chui_body = re.sub(r'^.{0,8}留意点[①②③０-９\d]?\s*[:：]?\s*', '', line).strip()
        chui_body = _clean_line(chui_body)
        if len(chui_body) < 10 or '留意する' not in chui_body:
            i += 1
            continue

        # 続く理由行を収集
        reasons = []
        j = i + 1
        while j < len(lines) and len(reasons) < 4:
            r_line = _clean_line(lines[j])
            # 次の留意点が来たら終了
            if (re.search(r'留意点[①②③]', r_line) or
                    ('留意する' in r_line and len(r_line) > 10 and not re.search(r'^.{0,4}理由', r_line))):
                break
            if re.search(r'^.{0,8}理由[①②③④]', r_line) and 'ため' in r_line:
                r_body = re.sub(r'^.{0,8}理由[①②③④]\s*[:：]?\s*', '', r_line).strip()
                r_body = _clean_line(r_body)
                if len(r_body) >= 10:
                    reasons.append(r_body)
            j += 1

        blocks.append({"留意点": chui_body, "理由": reasons})
        i = j if j > i + 1 else i + 1

    return blocks


def _analyze_chars(text: str) -> dict:
    """留意点・理由の文字数統計"""
    chui_lines = _extract_chui_lines(text)
    riyu_lines = _extract_riyu_lines(text)
    chui_lens = [len(c) for c in chui_lines]
    riyu_lens = [len(r) for r in riyu_lines]
    return {
        "chui_avg":     round(sum(chui_lens) / len(chui_lens)) if chui_lens else 0,
        "riyu_avg":     round(sum(riyu_lens) / len(riyu_lens)) if riyu_lens else 0,
        "chui_count":   len(chui_lens),
        "riyu_count":   len(riyu_lens),
        "chui_samples": chui_lens[:6],
        "riyu_samples": riyu_lens[:6],
    }


def _analyze_patterns(text: str) -> dict:
    """
    過去提案書のパターンを分析する。
    - 頻出する語尾表現
    - 頻出するキーワード（安全・工程・環境等の視点）
    - 留意点-理由ペアの構造
    """
    chui_lines = _extract_chui_lines(text)
    riyu_lines = _extract_riyu_lines(text)

    # 語尾パターン（留意点）
    chui_endings = []
    for c in chui_lines:
        m = re.search(r'(.{3,10}に留意する。?)$', c)
        if m:
            chui_endings.append(m.group(1))

    # 語尾パターン（理由）
    riyu_endings = []
    for r in riyu_lines:
        m = re.search(r'(.{4,15}ため。?)$', r)
        if m:
            riyu_endings.append(m.group(1))

    # 視点キーワード（理由行に含まれる視点語）
    axis_keywords = {
        "安全": ["安全", "事故", "リスク", "転落", "接触", "危険", "保護"],
        "工程": ["工程", "工期", "効率", "段取り", "スケジュール", "遅延", "短縮"],
        "環境": ["環境", "水質", "濁水", "騒音", "漁業", "観光", "住民", "生態"],
        "品質": ["品質", "精度", "確保", "完成", "コスト", "経済"],
    }
    axis_counts = {axis: 0 for axis in axis_keywords}
    for r in riyu_lines:
        for axis, kws in axis_keywords.items():
            if any(kw in r for kw in kws):
                axis_counts[axis] += 1

    return {
        "chui_endings": list(Counter(chui_endings).most_common(5)),
        "riyu_endings":  list(Counter(riyu_endings).most_common(5)),
        "axis_balance":  axis_counts,
        "paired_blocks": _extract_paired_blocks(text),
    }


# ──────────────────────────────────────────
# 公開API
# ──────────────────────────────────────────
def add(filename: str, file_bytes: bytes) -> str:
    """ファイルを保存してメタデータに追加。同名ファイルは上書き。"""
    _init()

    meta = _load_meta()
    duplicates = [m for m in meta if m["filename"] == filename]
    for dup in duplicates:
        fp = os.path.join(REF_DIR, dup["saved_as"])
        if os.path.exists(fp):
            os.remove(fp)
    meta = [m for m in meta if m["filename"] != filename]

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe = re.sub(r'[\\/:*?"<>|]', "_", filename)
    saved_as = f"{ts}_{safe}"
    file_path = os.path.join(REF_DIR, saved_as)

    with open(file_path, "wb") as f:
        f.write(file_bytes)

    text = extract_text(file_path, filename)
    char_count = _analyze_chars(text)
    patterns   = _analyze_patterns(text)

    ref_id = f"{ts}_{len(meta)}"
    txt_filename = f"{ref_id}.txt"
    txt_path = os.path.join(REF_DIR, txt_filename)
    with open(txt_path, "w", encoding="utf-8") as f:
        f.write(text)
    meta.append({
        "id":        ref_id,
        "filename":  filename,
        "saved_as":  saved_as,
        "added_at":  datetime.now().strftime("%Y-%m-%d %H:%M"),
        "chars":     char_count,
        "patterns":  {
            "chui_endings":  patterns["chui_endings"],
            "riyu_endings":  patterns["riyu_endings"],
            "axis_balance":  patterns["axis_balance"],
            "paired_count":  len(patterns["paired_blocks"]),
        },
        "text_file": txt_filename,
    })
    _save_meta(meta)
    return ref_id


def list_all() -> list:
    return _load_meta()


def delete(ref_id: str):
    meta = _load_meta()
    entry = next((m for m in meta if m["id"] == ref_id), None)
    if entry:
        fp = os.path.join(REF_DIR, entry["saved_as"])
        if os.path.exists(fp):
            os.remove(fp)
        txt_file = entry.get("text_file", "")
        if txt_file:
            txt_path = os.path.join(REF_DIR, txt_file)
            if os.path.exists(txt_path):
                os.remove(txt_path)
    _save_meta([m for m in meta if m["id"] != ref_id])


# ──────────────────────────────────────────
# AI向けコンテキスト生成
# ──────────────────────────────────────────
def _format_paired_blocks(blocks: list, max_blocks: int = 12) -> str:
    """留意点-理由ペアをAIが学習しやすい形式に整形"""
    if not blocks:
        return ""
    lines = []
    for i, b in enumerate(blocks[:max_blocks], 1):
        lines.append(f"  [{i}] 留意点：{b['留意点']}")
        for ri, r in enumerate(b["理由"], 1):
            lines.append(f"       理由{ri}：{r}")
    return "\n".join(lines)


def get_context() -> str:
    """
    全参照提案書のコンテキストをAIナレッジベースとして構築して返す。
    提案書が蓄積されるほど内容が充実し、AI生成精度が向上する。
    """
    meta = _load_meta()
    if not meta:
        return ""

    lines = []

    # ── 全件統計サマリー
    all_chui_avg = [m["chars"]["chui_avg"] for m in meta if m.get("chars", {}).get("chui_avg")]
    all_riyu_avg  = [m["chars"]["riyu_avg"]  for m in meta if m.get("chars", {}).get("riyu_avg")]
    total_chui = sum(m["chars"].get("chui_count", 0) for m in meta if m.get("chars"))
    total_riyu  = sum(m["chars"].get("riyu_count", 0) for m in meta if m.get("chars"))

    lines.append(f"▼ 蓄積ナレッジ統計（全{len(meta)}件の提案書）")
    if all_chui_avg:
        avg_c = round(sum(all_chui_avg) / len(all_chui_avg))
        lines.append(f"  留意点：累計{total_chui}件 / 平均{avg_c}字（{min(all_chui_avg)}〜{max(all_chui_avg)}字）")
    if all_riyu_avg:
        avg_r = round(sum(all_riyu_avg) / len(all_riyu_avg))
        lines.append(f"  理由  ：累計{total_riyu}件 / 平均{avg_r}字（{min(all_riyu_avg)}〜{max(all_riyu_avg)}字）")

    # ── 全件の視点バランス集計
    total_axis = Counter()
    for m in meta:
        ab = m.get("patterns", {}).get("axis_balance", {})
        for axis, cnt in ab.items():
            total_axis[axis] += cnt
    if total_axis:
        balance_str = "、".join(f"{k}:{v}件" for k, v in sorted(total_axis.items(), key=lambda x: -x[1]))
        lines.append(f"  理由視点バランス（全件）：{balance_str}")

    # ── 頻出語尾パターン（全件集計）
    all_chui_end = Counter()
    all_riyu_end  = Counter()
    for m in meta:
        for e, cnt in m.get("patterns", {}).get("chui_endings", []):
            all_chui_end[e] += cnt
        for e, cnt in m.get("patterns", {}).get("riyu_endings", []):
            all_riyu_end[e] += cnt
    if all_chui_end:
        top_chui = "、".join(f"「{e}」" for e, _ in all_chui_end.most_common(3))
        lines.append(f"  留意点の頻出語尾：{top_chui}")
    if all_riyu_end:
        top_riyu = "、".join(f"「{e}」" for e, _ in all_riyu_end.most_common(3))
        lines.append(f"  理由の頻出語尾  ：{top_riyu}")

    lines.append("")

    # ── 各提案書の詳細（留意点-理由ペア付き）
    for i, m in enumerate(meta, 1):
        ca = m.get("chars", {})
        pt = m.get("patterns", {})

        # ヘッダー
        stat = ""
        if ca.get("chui_avg"):
            stat = f"留意点≈{ca['chui_avg']}字・理由≈{ca['riyu_avg']}字"
        axis_str = ""
        if pt.get("axis_balance"):
            ab = pt["axis_balance"]
            axis_str = f"視点：安全{ab.get('安全',0)}/工程{ab.get('工程',0)}/環境{ab.get('環境',0)}/品質{ab.get('品質',0)}"
        lines.append(f"─── 提案書{i}：{m['filename']}  登録:{m['added_at']}  {stat}  {axis_str}")

        # 留意点-理由ペアブロック（最も重要な学習データ）
        text_file = m.get("text_file", "")
        if text_file:
            txt_path = os.path.join(REF_DIR, text_file)
            raw_text = open(txt_path, encoding="utf-8").read() if os.path.exists(txt_path) else ""
        else:
            raw_text = m.get("text", "")  # 旧形式との互換性
        if raw_text and not raw_text.startswith("["):
            blocks = _extract_paired_blocks(raw_text)
            if blocks:
                lines.append("  【留意点-理由ペア（学習用）】")
                lines.append(_format_paired_blocks(blocks, max_blocks=9))
            else:
                # ペア抽出できない場合は生テキストから留意点・理由行だけ抜粋
                chui_l = _extract_chui_lines(raw_text)
                riyu_l  = _extract_riyu_lines(raw_text)
                if chui_l:
                    lines.append("  【留意点】")
                    lines.extend(f"    ・{c}" for c in chui_l[:9])
                if riyu_l:
                    lines.append("  【理由】")
                    lines.extend(f"    ・{r}" for r in riyu_l[:12])
        lines.append("")

    return "\n".join(lines)


def extract_text_from_path(file_path: str) -> str:
    """ファイルパスからテキストを抽出する公開API。"""
    return extract_text(file_path, os.path.basename(file_path))
